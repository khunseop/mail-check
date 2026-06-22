"""
Mail Check 백엔드 (FastAPI)

Chrome 확장 프로그램이 새 메일 감지 시 POST /analyze 를 호출합니다.
첨부파일은 이미 ~/Downloads/{downloadFolder}/ 에 저장된 상태입니다.

실행:
    pip install fastapi uvicorn openpyxl
    uvicorn main:app --reload --port 8000
"""

import os
from pathlib import Path

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI(title="Mail Check Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Chrome 확장 → localhost 허용
    allow_methods=["POST", "GET"],
    allow_headers=["Content-Type"],
)

DOWNLOADS_DIR = Path.home() / "Downloads"


# ── 요청/응답 스키마 ──────────────────────────────────────

class AnalyzeRequest(BaseModel):
    subject: str
    body: str
    attachments: list[str] = []
    downloadFolder: str = ""


class AnalyzeResponse(BaseModel):
    replyText: str


# ── 유틸 ─────────────────────────────────────────────────

def resolve_attachment_paths(attachments: list[str], download_folder: str) -> list[Path]:
    """첨부파일 절대경로 목록 반환. 존재하는 파일만."""
    base = DOWNLOADS_DIR / download_folder if download_folder else DOWNLOADS_DIR
    return [base / name for name in attachments if (base / name).exists()]


def parse_excel(path: Path) -> list[dict]:
    """xlsx 파일을 dict 목록으로 반환."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 파싱 실패 ({path.name}): {e}")


# ── 엔드포인트 ────────────────────────────────────────────

@app.post("/analyze", response_model=AnalyzeResponse)
async def analyze(req: AnalyzeRequest):
    """
    메일 본문·첨부파일을 받아 처리 후 답장 텍스트 반환.

    - attachments: 파일명 목록 (경로 아님)
    - downloadFolder: Downloads 기준 상대 폴더명 (예: "mail-check")
    """
    paths = resolve_attachment_paths(req.attachments, req.downloadFolder)

    # Excel 파일 파싱 (있는 경우)
    excel_data: list[dict] = []
    for path in paths:
        if path.suffix.lower() in (".xlsx", ".xls"):
            excel_data = parse_excel(path)
            break  # 첫 번째 Excel만 처리 (필요 시 수정)

    # ── 여기에 실제 비즈니스 로직 작성 ──────────────────
    # 예시: excel_data를 가공해 reply_text 생성
    reply_text = build_reply(req.subject, req.body, excel_data)
    # ─────────────────────────────────────────────────────

    return AnalyzeResponse(replyText=reply_text)


@app.get("/health")
async def health():
    return {"status": "ok"}


# ── 답장 생성 로직 (커스터마이징 영역) ────────────────────

def build_reply(subject: str, body: str, excel_data: list[dict]) -> str:
    """
    메일 내용과 Excel 데이터를 바탕으로 답장 텍스트를 생성합니다.
    실제 비즈니스 로직에 맞게 수정하세요.
    """
    if not excel_data:
        return "안녕하세요. 메일 확인하였습니다. 검토 후 연락드리겠습니다."

    # 예시: 첫 번째 행 요약
    first_row = excel_data[0]
    summary = ", ".join(f"{k}: {v}" for k, v in first_row.items() if v is not None)
    return f"안녕하세요. 첨부파일 확인하였습니다.\n\n[확인 내용]\n{summary}\n\n검토 후 연락드리겠습니다."
